# -*- coding: utf-8 -*-
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import random,time

#read sheet data
def get_data(ro,col):
    work_b = load_workbook(filename='example.xlsx')  #here is the name of your xlsx file
    sheetnames =work_b.get_sheet_names()
    sheet = work_b.get_sheet_by_name(sheetnames[0])
    return sheet.cell(row  = ro, column = col).value

#random delay
def ti(cl):
    time.sleep(round(random.uniform(3, cl), 1))
    return

#compare string
def comp(str):
    str1 = 'attitude'
    tf = (str1 == str)
    return tf

#date adding
def add_date(date):
    year = int(date[0:4])
    month = int(date[4:6])
    day = int(date[6:8])
    if month == 2 :
        if year % 4 == 0:
            if day+6 > 29:
                day = day+6-29
                if month == 12:
                    month = 1
                    year = year + 1
                else:
                    month = month + 1
            else:
                day = day + 6
        if not(year % 4 == 0) :
            if day + 6 > 28:
                day = day + 6 - 28
                if month == 12:
                    month = 1
                    year = year + 1
                else:
                    month = month + 1
            else:
                day = day + 6
    elif month == 1 or 3 or 5 or 7 or 8 or 10 or 12:
        if day + 6 > 31:
            day = day + 6 - 31
            if month == 12:
                month = 1
                year = year + 1
            else:
                month = month + 1
        else:
            day = day + 6
    else:
        if day + 6 > 30:
            day = day + 6 - 30
            if month == 12:
                month = 1
                year = year + 1
            else:
                month = month + 1
        else:
            day = day + 6
    if month <10:
        m = '0'+str(month)
        if day <10:
            d = '0'+str(day)
            riqi = str(year) + m + d
        else:
            riqi = str(year) + m + str(day)
    else:
        if day < 10:
            d = '0' + str(day)
            riqi = str(year) + str(month) + d
        else:
            riqi = str(year) + str(month) + str(day)
    return riqi

def obtain_data(id, movie, start, end):
    wb = Workbook()  # Creat sheet
    ws = wb.active
    cookie = {"Cookie": "_T_WM=b1952cb717e779183c842ffb67ce77f5; ALF=1487696982; SCF=AtYH3r1ESsnUuFSSxzRrc7XVBo--4ru32oZQRNA9wbsaw23wqK94EImEZUrNjGaHk1HvE0tubC4WNpTPluxy62k.; SUB=_2A251gJn_DeRxGeRG6FMW8SrPwz6IHXVWiie3rDV6PUJbktBeLUrakW2V44EcbTwyzAOl9c6sapE6b89CYg..; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WhkWCJ.2GSIJhpkfJ1pHZ6Y5JpX5o2p5NHD95QE1hepS02Xe0nEWs4Dqcj4i--Ni-iWi-8si--NiK.4i-i2i--NiKLWiKnXi--Xi-isi-2pi--Xi-iFi-z7Ksv0KgHE9g4rUgH0; SUHB=0pHwqyOUkZ3oZo; SSOLoginState=1485105583"}
    headers = {"User-Agent":"Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36"}

    proxyHost = "proxy.abuyun.com"
    proxyPort = "9010"
    proxyUser = "H0068S5F6R5PI4UD"
    proxyPass = "A8F2A61C56E7B67B"
    proxyMeta = "http://%(user)s:%(pass)s@%(host)s:%(port)s" % {
        "host": proxyHost,
        "port": proxyPort,
        "user": proxyUser,
        "pass": proxyPass,
    }
    proxies = {
        "http": proxyMeta,
        "https": proxyMeta,
    }

    zong = False
    ci = 1
    cc = 1
    while(not(zong)):
        try:
            print('Find total number. Round: '+str(ci))
            ci = ci+1
            if ci == 5:
                ws['A1'] = id
                ws['B1'] = movie
                ws['C1'] = '共0条'
                wb.save(id + ".xlsx")
                return  cc
            url = 'http://weibo.cn/search/mblog?hideSearchFrame=&keyword=' + movie + '&advancedfilter=1&starttime=' + start + '&endtime=' + end + '&sort=time&page=1'
            html = requests.get(url, cookies=cookie, headers = headers,proxies=proxies, timeout=10)
            cc = html.status_code
            ti(4)
            soup = BeautifulSoup(html.text, "html.parser")
            # Record total number
            total_num = soup.find('span', {'class': 'cmt'})
            print(total_num.string)
            ws['A1'] = id
            ws['B1'] = movie
            ws['C1'] = total_num.string
            zong = True
        except:
            zong = False

    j = 1
    k = 1
    # divide 2 month into 10 parts and do for loop
    for p in range(1, 10+1):
        try:
            end = add_date(start)
            k = k+1
            j = k
            ws['A' + str(k)] = p
            url = 'http://weibo.cn/search/mblog?hideSearchFrame=&keyword=' + movie + '&advancedfilter=1&starttime=' + start + '&endtime=' + end + '&sort=time&page=1'
            html_1 = requests.get(url,cookies=cookie,proxies=proxies,timeout = 10)
            ti(4)
            print(html_1.status_code)
            soup_1 = BeautifulSoup(html_1.text, "html.parser")
            total_num_1 = soup_1.find('span', {'class': 'cmt'})
            print(total_num_1.string)
            ws['B' + str(k)] = total_num_1.string
            # Record information of each weibo in each time buckets search results Page 1
            for content in soup_1.find_all('span', {'class': 'ctt'}):
                k = k + 1
                print(content)
                ws['G' + str(k)] = content.text
                # html for 'likes'
                a_1 = content.find_next("a")
                http = a_1.get('href')
                while (not (comp(http[16:24]))):
                    a_next = a_1.find_next("a")
                    http = a_next.get('href')
                    a_1 = a_next
                # 'likes' read
                ws['C' + str(k)] = a_1.string
                a_2 = a_1.find_next("a")
                # retweets
                ws['D' + str(k)] = a_2.string
                a_3 = a_2.find_next("a")
                # comment
                ws['E' + str(k)] = a_3.string
                time = a_3.find_next("span")
                # time
                # print (time.string[0:19])
                ws['F' + str(k)] = time.string

            # Record account name and fan amount
            for name in soup_1.find_all('a', {'class': 'nk'}):
                j = j + 1
                # Users' name
                ws['A' + str(j)] = name.string
                print(name.string)
                href = name.get('href')
                # fans
                ws['B' + str(j)] = href

            start = add_date(start)
            # output
            wb.save(id + ".xlsx")
        except Exception:
            start = add_date(start)
            wb.save(id + ".xlsx")
            continue
    return code

#main
for ro in range(124,288+1):
    try:
        id = str(get_data(ro,1))
        movie = str(get_data(ro,2))
        start = str(get_data(ro,3))
        end = str(get_data(ro,4))
        print(id,movie,start,end)
        code =obtain_data(id, movie, start, end)
        print(code)
    except Exception:
        continue
